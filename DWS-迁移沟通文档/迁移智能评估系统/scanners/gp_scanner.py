"""Greenplum 源端扫描器

支持从调研模板Excel解析或从数据库直连采集元数据
"""

import os
import re
from core.models import MigrationMetadata
from scanners.base import BaseScanner


class GPScanner(BaseScanner):
    """Greenplum 扫描器"""

    def __init__(self, excel_path: str = ""):
        super().__init__(source_type="gp")
        self.excel_path = excel_path

    def scan(self) -> MigrationMetadata:
        """执行扫描"""
        if self.excel_path and os.path.exists(self.excel_path):
            return self._parse_excel(self.excel_path)
        return self._sample_data()

    def _parse_excel(self, filepath: str) -> MigrationMetadata:
        """从调研模板Excel解析"""
        meta = MigrationMetadata()
        try:
            import openpyxl
        except ImportError:
            print("[WARN] 需要 openpyxl: pip install openpyxl")
            return meta

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if not any(v.strip() for v in vals):
                continue

            # 尝试匹配关键字段
            text = " ".join(v.strip() for v in vals if v.strip())

            if "GP" in text.upper() or "Greenplum" in text:
                for v in vals:
                    if "GP" in str(v).upper():
                        meta.db_version = str(v).strip()
                        break
            if "集群规模" in text or ("集群" in text and "规模" in text):
                meta.cluster_scale = vals[2].strip() if len(vals) > 2 else ""
            if "容量" in text and "TB" in text:
                for v in vals:
                    if "TB" in str(v):
                        meta.total_capacity = str(v).strip()
                        break
            if "表" in text and ("视图" in text or "函数" in text):
                nums = re.findall(r'\d+', text)
                if len(nums) >= 1:
                    meta.table_count = int(nums[0])
                if len(nums) >= 2:
                    meta.view_count = int(nums[1])
                if len(nums) >= 3:
                    meta.function_count = int(nums[2])

            if "Kettle" in text:
                meta.etl_tool = "Kettle"
                for v in vals:
                    nums = re.findall(r'\d+', v)
                    if nums:
                        meta.etl_task_count = int(nums[0])

            if "TaskCTL" in text or "调度" in text:
                for v in vals:
                    if "TaskCTL" in str(v):
                        meta.scheduler_tool = str(v).strip()

        wb.close()
        return meta

    @staticmethod
    def _sample_data() -> MigrationMetadata:
        """返回典型GP客户样本数据"""
        return MigrationMetadata(
            db_version="Greenplum 5.23",
            kernel_version="PostgreSQL 8.3.23",
            cluster_scale="2 Master + 6 Segment",
            hardware_info="Master: 2*16Core 64G; Segment: 2*20Core 128G 18*3.84T SSD",
            total_capacity="35TB",
            table_count=2700,
            view_count=1100,
            function_count=1300,
            sequence_count=350,
            schema_count=45,
            user_count=120,
            database_count=8,
            partition_table_count=850,
            compression_table_count=1200,
            has_realtime_data=True,
            udf_languages={
                "plpgsql": 1027,
                "plpythonu": 18,
                "plperl": 247,
                "plperlu": 8,
            },
            data_types_used=["INT", "BIGINT", "NUMERIC", "VARCHAR", "TEXT",
                             "TIMESTAMP", "DATE", "BOOLEAN", "JSONB", "SERIAL"],
            etl_tool="Kettle 7.x",
            etl_task_count=1300,
            source_db_types="Oracle, MySQL, 文件系统",
            source_system_count=10,
            scheduler_tool="iControl",
            scheduler_task_count=1500,
            scheduler_frequency="每日凌晨15分钟批量",
            api_tool="XXX数据服务V1.0",
            api_count=200,
            app_count=2,
            bi_tools=["FineReport 9.x"],
            # 补充字段
            security_model="PostgreSQL RBAC (与DWS兼容)",
            encryption_method="pgcrypto扩展",
            audit_capability="GP审计日志",
            tls_version="TLS 1.2",
            uses_filesystem_access=False,
            uses_lbac=False,
            has_superuser_privileges=True,
            source_charset="UTF-8",
            has_multibyte_chars=True,
            collation_name="默认",
            jdbc_driver_version="PostgreSQL JDBC 42.x",
            connection_pool="HikariCP",
            orm_framework="MyBatis",
            app_language="Java",
            default_isolation_level="READ COMMITTED",
            supports_xa=False,
            lock_timeout_seconds=0,
            has_long_transactions=True,
            uses_autonomous_transaction=False,
            cdc_tool="无原生CDC",
            has_cdc_active=False,
            daily_data_growth_gb=30.0,
            max_table_size_gb=1000.0,
            has_data_skew=True,
            workload_type="OLAP",
            peak_concurrent_queries=80,
            avg_query_response_seconds=5.0,
        )
